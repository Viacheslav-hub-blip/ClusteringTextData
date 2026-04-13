"""Excel loader for the agentic banking clustering project."""

from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook


def load_comments_from_excel(
    path: str | Path,
    *,
    text_column: str = "comment",
    id_column: str = "comment_id",
    sheet_name: str | None = None,
    limit: int | None = None,
) -> list[dict]:
    """Load comments from an Excel file."""
    workbook = load_workbook(filename=Path(path), read_only=True, data_only=True)
    worksheet = workbook[sheet_name] if sheet_name else workbook.active
    rows = worksheet.iter_rows(values_only=True)
    header = next(rows)
    if header is None:
        return []

    header_map = {str(value).strip(): index for index, value in enumerate(header) if value is not None}
    if text_column not in header_map:
        raise KeyError(f"Column '{text_column}' was not found in {path}")

    text_index = header_map[text_column]
    id_index = header_map.get(id_column)
    comments: list[dict] = []

    for row_number, row in enumerate(rows, start=2):
        text = row[text_index] if text_index < len(row) else None
        if text is None:
            continue
        comment_id = row[id_index] if id_index is not None and id_index < len(row) else row_number - 1
        comments.append({"comment_id": str(comment_id), "text": str(text)})
        if limit is not None and len(comments) >= limit:
            break

    return comments
