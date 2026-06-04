"""Простой API для запуска кластеризации из Jupyter Notebook или IDE.

Файл содержит:
- ``cluster_text_data`` — запуск кластеризации по пути к файлу, DataFrame или списку словарей;
- ``load_text_rows`` — загрузка строк из DataFrame, CSV, XLSX или списка словарей;
- ``append_cluster_labels`` — добавление результатов кластеризации к исходным строкам;
- ``save_rows`` — сохранение результата в CSV или XLSX;
- ``_rows_from_dataframe`` — преобразование DataFrame в список словарей;
- ``_rows_from_records`` — подготовка списка словарей;
- ``_read_csv_rows`` — чтение CSV-файла;
- ``_read_xlsx_rows`` — чтение XLSX-файла;
- ``_to_dataframe_if_possible`` — преобразование результата в DataFrame при наличии pandas;
- ``_write_csv_rows`` — запись CSV-файла;
- ``_write_xlsx_rows`` — запись XLSX-файла.
"""

from __future__ import annotations

import csv
from collections.abc import Mapping, Sequence
from os import PathLike
from pathlib import Path
from typing import Any

from langchain_core.embeddings import Embeddings
from langchain_core.language_models import BaseChatModel


def cluster_text_data(
        data: str | PathLike[str] | Sequence[Mapping[str, Any]] | Any,
        *,
        llm: BaseChatModel,
        embeddings: Embeddings,
        text_column: str = "text",
        id_column: str | None = None,
        sheet_name: str | int | None = 0,
        output_path: str | PathLike[str] | None = None,
        return_dataframe: bool = True,
        include_details: bool = False,
        **pipeline_kwargs: Any,
) -> Any:
    """Кластеризует комментарии из файла, DataFrame или списка словарей.

    Args:
        data: Путь к CSV/XLSX, pandas DataFrame или список словарей с комментариями.
        llm: Готовая LangChain chat-модель для выбора группы и опционального нейминга.
        embeddings: Готовая LangChain embedding-модель для поиска похожих комментариев.
        text_column: Название колонки с текстом комментария.
        id_column: Название колонки с ID комментария. Если ``None``, используется
            ``comment_id`` или порядковый номер строки.
        sheet_name: Имя или индекс листа XLSX-файла.
        output_path: Путь для сохранения результата в CSV или XLSX. Если ``None``,
            файл не сохраняется.
        return_dataframe: Если ``True`` и установлен pandas, возвращает DataFrame.
        include_details: Если ``True``, добавляет технические поля решения LLM.
        pipeline_kwargs: Дополнительные параметры ``VectorLLMClusteringPipeline``.

    Returns:
        DataFrame при доступном pandas и ``return_dataframe=True``, иначе список словарей
        с исходными колонками и добавленным ``group_name``.
    """
    from .pipelines import VectorLLMClusteringPipeline

    rows = load_text_rows(
        data,
        text_column=text_column,
        id_column=id_column,
        sheet_name=sheet_name,
    )
    raw_comments = [
        {
            "comment_id": row["comment_id"],
            "text": row[text_column],
        }
        for row in rows
    ]

    pipeline = VectorLLMClusteringPipeline(
        llm=llm,
        embeddings=embeddings,
        **pipeline_kwargs,
    )
    result = pipeline.run(raw_comments)
    output_rows = append_cluster_labels(rows, result, include_details=include_details)

    if output_path is not None:
        save_rows(output_rows, output_path)

    if return_dataframe:
        return _to_dataframe_if_possible(output_rows)
    return output_rows


def load_text_rows(
        data: str | PathLike[str] | Sequence[Mapping[str, Any]] | Any,
        *,
        text_column: str = "text",
        id_column: str | None = None,
        sheet_name: str | int | None = 0,
) -> list[dict[str, Any]]:
    """Загружает строки с комментариями из поддерживаемого источника.

    Args:
        data: Путь к CSV/XLSX, pandas DataFrame или список словарей.
        text_column: Название поля с текстом комментария.
        id_column: Название поля с ID комментария.
        sheet_name: Имя или индекс листа XLSX-файла.

    Returns:
        Список словарей, где у каждой строки есть ``comment_id`` и поле с текстом.

    Raises:
        ValueError: Если источник пустой, формат файла не поддержан или нет текстовой колонки.
    """
    if isinstance(data, (str, PathLike)):
        path = Path(data)
        suffix = path.suffix.lower()
        if suffix == ".csv":
            rows = _read_csv_rows(path)
        elif suffix in {".xlsx", ".xlsm"}:
            rows = _read_xlsx_rows(path, sheet_name=sheet_name)
        else:
            raise ValueError(f"Поддерживаются только CSV и XLSX-файлы: {path}")
        return _rows_from_records(rows, text_column=text_column, id_column=id_column)

    if hasattr(data, "to_dict") and hasattr(data, "columns"):
        return _rows_from_dataframe(data, text_column=text_column, id_column=id_column)

    if isinstance(data, Sequence) and not isinstance(data, (str, bytes, bytearray)):
        return _rows_from_records(data, text_column=text_column, id_column=id_column)

    raise ValueError("Источник должен быть путем к файлу, DataFrame или списком словарей.")


def append_cluster_labels(
        rows: Sequence[Mapping[str, Any]],
        clustering_result: Mapping[str, Any],
        *,
        include_details: bool = False,
) -> list[dict[str, Any]]:
    """Добавляет к исходным строкам названия групп из результата pipeline.

    Args:
        rows: Исходные строки с полем ``comment_id``.
        clustering_result: Результат ``VectorLLMClusteringPipeline.run``.
        include_details: Если ``True``, добавляет ``decision_type``, ``decision_reason``
            и ``normalized_text``.

    Returns:
        Список словарей с добавленным полем ``group_name``.
    """
    comments = clustering_result.get("comments", [])
    by_id = {str(comment.get("comment_id", "")): comment for comment in comments}
    output_rows: list[dict[str, Any]] = []

    for row in rows:
        output_row = dict(row)
        comment = by_id.get(str(output_row.get("comment_id", "")), {})
        output_row["group_name"] = comment.get("group_name", "")
        if include_details:
            output_row["normalized_text"] = comment.get("normalized_text", "")
            output_row["decision_type"] = comment.get("decision_type", "")
            output_row["decision_reason"] = comment.get("decision_reason", "")
        output_rows.append(output_row)

    return output_rows


def save_rows(rows: Sequence[Mapping[str, Any]], output_path: str | PathLike[str]) -> None:
    """Сохраняет строки результата в CSV или XLSX.

    Args:
        rows: Строки результата для сохранения.
        output_path: Путь к выходному CSV или XLSX-файлу.

    Returns:
        ``None``.

    Raises:
        ValueError: Если расширение файла не поддерживается.
    """
    path = Path(output_path)
    suffix = path.suffix.lower()
    if suffix == ".csv":
        _write_csv_rows(rows, path)
        return
    if suffix in {".xlsx", ".xlsm"}:
        _write_xlsx_rows(rows, path)
        return
    raise ValueError(f"Поддерживаются только CSV и XLSX-файлы: {path}")


def _rows_from_dataframe(dataframe: Any, *, text_column: str, id_column: str | None) -> list[dict[str, Any]]:
    """Преобразует pandas DataFrame в подготовленный список строк.

    Args:
        dataframe: DataFrame с комментариями.
        text_column: Название колонки с текстом.
        id_column: Название колонки с ID.

    Returns:
        Список словарей с обязательными полями ``comment_id`` и ``text_column``.
    """
    return _rows_from_records(
        dataframe.to_dict(orient="records"),
        text_column=text_column,
        id_column=id_column,
    )


def _rows_from_records(
        records: Sequence[Mapping[str, Any]],
        *,
        text_column: str,
        id_column: str | None,
) -> list[dict[str, Any]]:
    """Готовит список словарей к передаче в pipeline.

    Args:
        records: Исходные записи.
        text_column: Название поля с текстом.
        id_column: Название поля с ID.

    Returns:
        Список строк с непустым ``comment_id`` и текстовой колонкой.

    Raises:
        ValueError: Если нет текстовой колонки или нет строк.
    """
    rows: list[dict[str, Any]] = []
    source_id_column = id_column or "comment_id"

    for index, record in enumerate(records, start=1):
        row = dict(record)
        if text_column not in row:
            raise ValueError(f"В данных нет текстовой колонки: {text_column}")
        text = "" if row[text_column] is None else str(row[text_column]).strip()
        if not text:
            continue
        comment_id = str(row.get(source_id_column, "")).strip() or str(index)
        row["comment_id"] = comment_id
        row[text_column] = text
        rows.append(row)

    if not rows:
        raise ValueError("Нет строк с непустым текстом комментария.")
    return rows


def _read_csv_rows(path: Path) -> list[dict[str, Any]]:
    """Читает CSV-файл в список словарей.

    Args:
        path: Путь к CSV-файлу.

    Returns:
        Список строк из CSV-файла.
    """
    last_error: UnicodeDecodeError | None = None
    for encoding in ("utf-8-sig", "utf-8", "cp1251"):
        try:
            with path.open("r", encoding=encoding, newline="") as file:
                return list(csv.DictReader(file))
        except UnicodeDecodeError as exc:
            last_error = exc
    if last_error is not None:
        raise last_error
    return []


def _read_xlsx_rows(path: Path, *, sheet_name: str | int | None) -> list[dict[str, Any]]:
    """Читает XLSX-файл в список словарей.

    Args:
        path: Путь к XLSX-файлу.
        sheet_name: Имя или индекс листа. Если ``None``, используется активный лист.

    Returns:
        Список строк из XLSX-файла.
    """
    from openpyxl import load_workbook

    workbook = load_workbook(path, read_only=True, data_only=True)
    if sheet_name is None:
        sheet = workbook.active
    elif isinstance(sheet_name, int):
        sheet = workbook.worksheets[sheet_name]
    else:
        sheet = workbook[sheet_name]

    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []

    headers = [str(value).strip() if value is not None else "" for value in rows[0]]
    result: list[dict[str, Any]] = []
    for values in rows[1:]:
        result.append({
            header: value
            for header, value in zip(headers, values, strict=False)
            if header
        })
    return result


def _to_dataframe_if_possible(rows: Sequence[Mapping[str, Any]]) -> Any:
    """Преобразует строки в pandas DataFrame, если pandas установлен.

    Args:
        rows: Строки результата.

    Returns:
        DataFrame при доступном pandas, иначе список словарей.
    """
    try:
        import pandas as pd
    except ImportError:
        return list(rows)
    return pd.DataFrame(list(rows))


def _write_csv_rows(rows: Sequence[Mapping[str, Any]], path: Path) -> None:
    """Записывает строки в CSV-файл.

    Args:
        rows: Строки для записи.
        path: Путь к CSV-файлу.

    Returns:
        ``None``.
    """
    rows = list(rows)
    path.parent.mkdir(parents=True, exist_ok=True)
    fieldnames = list(rows[0].keys()) if rows else []
    with path.open("w", encoding="utf-8-sig", newline="") as file:
        writer = csv.DictWriter(file, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def _write_xlsx_rows(rows: Sequence[Mapping[str, Any]], path: Path) -> None:
    """Записывает строки в XLSX-файл.

    Args:
        rows: Строки для записи.
        path: Путь к XLSX-файлу.

    Returns:
        ``None``.
    """
    from openpyxl import Workbook

    rows = list(rows)
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "clusters"

    headers = list(rows[0].keys()) if rows else []
    sheet.append(headers)
    for row in rows:
        sheet.append([row.get(header, "") for header in headers])

    workbook.save(path)
